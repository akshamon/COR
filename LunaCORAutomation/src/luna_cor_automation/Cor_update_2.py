from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.keys import Keys
import time
from pywinauto.keyboard import send_keys
import pandas as pd
from colorama import Fore, Back, Style
from getpass import getpass
import openpyxl
import cv2
import os

wb= openpyxl.load_workbook('Policy.xlsx')
sheet_obj = wb.active
 #Give the account credentials
email=input("Enter the email id")
password= getpass('Please enter your account password: ')
Game_name='//*[@id="game_tile_amzn1.adg.product.090330b3-1fe1-49a8-99e7-8b24f5f55227"]/div/div/div/div/img'
#give the game title 
#Game_name='//img[@alt="Watch Dogs 2"]'
#Game_name='//*[@id="game_tile_amzn1.adg.product.7a41b13f-f236-4830-bf53-1824ff552c5f"]'
game_wait=120
for x in range(1, sheet_obj.max_row):
    name_of_router = sheet_obj.cell(x+1,1).value
    print(name_of_router)
# scan available Wifi networks
        #os.system('cmd /c "netsh wlan show networks"')

    # input Wifi name
    #If you're not yet connected, try connecting to a previously connected SSID again!
    #name_of_router = networks[networks['Network_name']!='']

    #print(networks[networks['Network_name']!='']

    # connect to the given wifi network
    os.system(f'''cmd /c "netsh wlan connect name={name_of_router}"''')

    fake_controller = """
        window.navigator.getGamepads = () =>
            [
                {
                    axes: [0.01, 0.01, 0.02, 0.04],
                    buttons: [
                        { pressed: false, value: 0 },
                        { pressed: false, value: 0 },
                        { pressed: false, value: 0 },
                        { pressed: false, value: 0 },
                        { pressed: false, value: 0 },
                        { pressed: false, value: 0 },
                        { pressed: false, value: 0 },
                        { pressed: false, value: 0 },
                        { pressed: false, value: 0 },
                        { pressed: false, value: 0 },
                        { pressed: false, value: 0 },
                        { pressed: false, value: 0 },
                        { pressed: false, value: 0 },
                        { pressed: false, value: 0 },
                        { pressed: false, value: 0 },
                        { pressed: false, value: 0 },
                        { pressed: false, value: 0 },
                        { pressed: false, value: 0 }
                    ],
                    connected: true,
                    id: 'Xbox Wireless Controller (STANDARD GAMEPAD Vendor: 045e Product: 02fd)',
                    index: 0,
                    mapping: 'standard',
                    timestamp: 177550
                }
            ]
    """
    #Give the account credentials
    
        #give the game title 
        #Game_name='//img[@alt="Watch Dogs 2"]'
        #Game_name='//*[@id="game_tile_amzn1.adg.product.7a41b13f-f236-4830-bf53-1824ff552c5f"]'
    driver_luna = webdriver.Chrome()
    driver_luna.get("https://luna.amazon.com/settings")
    driver_luna.find_element(By.ID, "ap_email").send_keys(email)
    driver_luna.find_element(By.ID,"ap_password").send_keys(password)
    driver_luna.find_element(By.ID, "signInSubmit").click()
    driver_luna.implicitly_wait(20)
    time.sleep(3)
    WebDriverWait(driver_luna,120).until(EC.presence_of_element_located((By.ID, "collection_sub_nav_settings")))
    driver_luna.find_element(By.XPATH,value='//*[@id="collection_sub_nav_settings"]/div/nav/ol/div[7]/div').click()
    driver_luna.find_element(By.ID,'react-select-2-input').send_keys(Keys.ARROW_DOWN)
    driver_luna.find_element(By.ID,'react-select-3-input').send_keys(Keys.ENTER,Keys.ENTER)
    driver_luna.save_screenshot(r'C:\Users\prebyada\Desktop\Screenshots\English.png')
    time.sleep(2)
    driver_luna.find_element(By.ID,'item_nav_library').click()
    WebDriverWait(driver_luna,120).until(EC.presence_of_element_located((By.ID, "collection_library_certifier")))
    driver_luna.implicitly_wait(240)
    driver_luna.execute_script(fake_controller)
    driver_luna.find_element(By.XPATH,value=Game_name).click()
    driver_luna.implicitly_wait(60)
    time.sleep(5)
    driver_luna.find_element(By.ID ,"start_game_session__action").click()
    driver_luna.fullscreen_window()
    time.sleep(3)
    WebDriverWait(driver_luna,120).until(EC.presence_of_element_located((By.ID, "abandon_button")))
    #checks for exit button
    if driver_luna.find_element(By.XPATH,value='//*[@id="abandon_button"]/div/div'):
        print(Fore.GREEN+"Game launching for COR-US")
    elif driver_luna.find_element(By.XPATH,value='//*[@id="item_primary_game_detail_button"]'):
        print(Fore.RED+"Error launching game")
        driver_luna.save_screenshot(r'C:\Users\prebyada\Desktop\Screenshots\Error_in_launching_English_Us.png')
    else:
        print(Fore.RED+"Error launching game")
        driver_luna.save_screenshot(r'C:\Users\prebyada\Desktop\Screenshots\Error_in_launching_English_Us.png')
        #increase or decrease the game time based on the launch time
    time.sleep(game_wait)
        #change location of drive
    driver_luna.save_screenshot(r'C:\Users\prebyada\Desktop\Screenshots\OverideEnglish_1.png')
    image = cv2.imread("C:\\Users\\prebyada\\Desktop\\Screenshots\\OverideEnglish_1.png", 0)
    if cv2.countNonZero(image) == 0:
        print("Image is black")
    else:
        print("Colored image")
    send_keys('+{TAB}')
    time.sleep(10)
    send_keys("{VK_DOWN}")
    time.sleep(4)
    send_keys("{ENTER}")
    time.sleep(4)
    send_keys("{ENTER}")
    time.sleep(8)
    WebDriverWait(driver_luna,120).until(EC.presence_of_element_located((By.ID, "gameplay_survey_skip")))
    driver_luna.find_element(By.XPATH,value='//*[@id="gameplay_survey_skip"]/div/div').click()
    time.sleep(3)
    break
for x in range(2, sheet_obj.max_row):
    name_of_router = sheet_obj.cell(x+1,1).value
    print(name_of_router)
    os.system(f'''cmd /c "netsh wlan connect name={name_of_router}"''')


        #Canada
    driver_luna.get("https://luna.amazon.com/settings")
    time.sleep(2)
    WebDriverWait(driver_luna,120).until(EC.presence_of_element_located((By.ID, "collection_sub_nav_settings")))
    driver_luna.find_element(By.XPATH,value='//*[@id="collection_sub_nav_settings"]/div/nav/ol/div[7]/div').click()
    time.sleep(2)
    WebDriverWait(driver_luna,120).until(EC.presence_of_element_located((By.ID, "react-select-2-input")))
    driver_luna.find_element(By.ID,'react-select-2-input').send_keys(Keys.ARROW_DOWN)
    driver_luna.find_element(By.ID,'react-select-3-input').send_keys(Keys.ENTER,Keys.ARROW_DOWN,Keys.ENTER)
    driver_luna.save_screenshot(r'C:\Users\prebyada\Desktop\Screenshots\Canada.png')
    time.sleep(2)
    driver_luna.find_element(By.ID,'item_nav_library').click()
    WebDriverWait(driver_luna,120).until(EC.presence_of_element_located((By.ID, "collection_library_certifier")))
    driver_luna.implicitly_wait(240)
    driver_luna.execute_script(fake_controller)
    driver_luna.find_element(By.XPATH,value=Game_name).click()
    #luna wait time 
    time.sleep(30)
    driver_luna.find_element(By.ID ,"start_game_session__action").click()
    driver_luna.fullscreen_window()
    time.sleep(3)
    WebDriverWait(driver_luna,120).until(EC.presence_of_element_located((By.ID, "abandon_button")))
    if driver_luna.find_element(By.XPATH,value='//*[@id="abandon_button"]/div/div'):
        print(Fore.GREEN+"Game launching for COR-Canada")
    elif driver_luna.find_element(By.XPATH,value='//*[@id="item_primary_game_detail_button"]'):
        print(Fore.RED+"Error launching game")
        driver_luna.save_screenshot(r'C:\Users\prebyada\Desktop\Screenshots\Error_in_launching_Canada.png')
    else:
        print(Fore.RED+"Error launching game")
        driver_luna.save_screenshot(r'C:\Users\prebyada\Desktop\Screenshots\Error_in_launching_Canada.png')
        #increase launch time below
    time.sleep(game_wait)
        #change location of drive
    driver_luna.save_screenshot(r'C:\Users\prebyada\Desktop\Screenshots\OverideCanada_1.png')
    image = cv2.imread("C:\\Users\\prebyada\\Desktop\\Screenshots\\OverideCanada_1.png", 0)
    if cv2.countNonZero(image) == 0:
        print("Image is black")
    else:
        print("Colored image")
    send_keys('+{TAB}')
    time.sleep(10)
    send_keys("{VK_DOWN}")
    time.sleep(4)
    send_keys("{ENTER}")
    time.sleep(4)
    send_keys("{ENTER}")
    time.sleep(8)
    driver_luna.find_element(By.XPATH,value='//*[@id="gameplay_survey_skip"]/div/div').click()
    time.sleep(3)
    break

for x in range(3, sheet_obj.max_row):
    name_of_router = sheet_obj.cell(x+1,1).value
    print(name_of_router)
    os.system(f'''cmd /c "netsh wlan connect name={name_of_router}"''')


        #Great Britian
    driver_luna.get("https://luna.amazon.com/settings")
    time.sleep(2)
    WebDriverWait(driver_luna,120).until(EC.presence_of_element_located((By.ID, "collection_sub_nav_settings")))
    driver_luna.find_element(By.XPATH,value='//*[@id="collection_sub_nav_settings"]/div/nav/ol/div[7]/div').click()
    time.sleep(2)
    WebDriverWait(driver_luna,120).until(EC.presence_of_element_located((By.ID, "react-select-2-input")))
    driver_luna.find_element(By.ID,'react-select-2-input').send_keys(Keys.ARROW_DOWN)
    driver_luna.find_element(By.ID,'react-select-3-input').send_keys(Keys.ENTER,Keys.ARROW_DOWN,Keys.ARROW_DOWN,Keys.ENTER)
    driver_luna.save_screenshot(r'C:\Users\prebyada\Desktop\Screenshots\GB.png')
    time.sleep(2)
    driver_luna.find_element(By.ID,'item_nav_library').click()
    WebDriverWait(driver_luna,120).until(EC.presence_of_element_located((By.ID, "collection_library_certifier")))
    driver_luna.implicitly_wait(240)
    driver_luna.execute_script(fake_controller)
    driver_luna.find_element(By.XPATH,value=Game_name).click()
    driver_luna.implicitly_wait(60)
        #luna wait time
    time.sleep(30)
    driver_luna.find_element(By.ID ,"start_game_session__action").click()
    driver_luna.fullscreen_window()
    time.sleep(3)
    WebDriverWait(driver_luna,120).until(EC.presence_of_element_located((By.ID, "abandon_button")))
    if driver_luna.find_element(By.XPATH,value='//*[@id="abandon_button"]/div/div'):
        print(Fore.GREEN+"Game launching for COR-GB")
    elif driver_luna.find_element(By.XPATH,value='//*[@id="item_primary_game_detail_button"]'):
        print(Fore.RED+"Error launching game")
        driver_luna.save_screenshot(r'C:\Users\prebyada\Desktop\Screenshots\Error_in_launching_GB.png')
    else:
        print(Fore.RED+"Error launching game")
        driver_luna.save_screenshot(r'C:\Users\prebyada\Desktop\Screenshots\Error_in_launching_GB.png')
        #increase/decrease launch time below
    time.sleep(game_wait)
        #change location of drive
    driver_luna.save_screenshot(r'C:\Users\prebyada\Desktop\Screenshots\OverideGB_1.png')
    image = cv2.imread("C:\\Users\\prebyada\\Desktop\\Screenshots\\OverideGB_1.png", 0)
    if cv2.countNonZero(image) == 0:
        print("Image is black")
    else:
        print("Colored image")
    send_keys('+{TAB}')
    time.sleep(10)
    send_keys("{VK_DOWN}")
    time.sleep(4)
    send_keys("{ENTER}")
    time.sleep(4)
    send_keys("{ENTER}")
    time.sleep(8)
    driver_luna.find_element(By.XPATH,value='//*[@id="gameplay_survey_skip"]/div/div').click()
    time.sleep(3)
    break

for x in range(4, sheet_obj.max_row):
    name_of_router = sheet_obj.cell(x+1,1).value
    print(name_of_router) 
    os.system(f'''cmd /c "netsh wlan connect name={name_of_router}"''')
  
        #Germany
    driver_luna.get("https://luna.amazon.com/settings")
    time.sleep(2)
    WebDriverWait(driver_luna,120).until(EC.presence_of_element_located((By.ID, "collection_sub_nav_settings")))
    driver_luna.find_element(By.XPATH,value='//*[@id="collection_sub_nav_settings"]/div/nav/ol/div[7]/div').click()
    time.sleep(2)
    WebDriverWait(driver_luna,120).until(EC.presence_of_element_located((By.ID, "react-select-2-input")))
    driver_luna.find_element(By.ID,'react-select-2-input').send_keys(Keys.ARROW_DOWN)
    driver_luna.find_element(By.ID,'react-select-3-input').send_keys(Keys.ENTER,Keys.ARROW_DOWN,Keys.ARROW_DOWN,Keys.ARROW_DOWN,Keys.ENTER)
    driver_luna.save_screenshot(r'C:\Users\prebyada\Desktop\Screenshots\Germany.png')
    time.sleep(2)
    driver_luna.find_element(By.ID,'item_nav_library').click()
    WebDriverWait(driver_luna,120).until(EC.presence_of_element_located((By.ID, "collection_library_certifier")))
    driver_luna.implicitly_wait(240)
    driver_luna.execute_script(fake_controller)
    driver_luna.find_element(By.XPATH,value=Game_name).click()
    driver_luna.implicitly_wait(60)
        #Luna wait time 
    time.sleep(30)
    driver_luna.find_element(By.ID ,"start_game_session__action").click()
    driver_luna.fullscreen_window()
    time.sleep(3)
    WebDriverWait(driver_luna,120).until(EC.presence_of_element_located((By.ID, "abandon_button")))
    if driver_luna.find_element(By.XPATH,value='//*[@id="abandon_button"]/div/div'):
        print(Fore.GREEN+"Game launching for COR-Germany")
    elif driver_luna.find_element(By.XPATH,value='//*[@id="item_primary_game_detail_button"]'):
        print(Fore.RED+"Error launching game")
        driver_luna.save_screenshot(r'C:\Users\prebyada\Desktop\Screenshots\Error_in_launching_Germany.png')
    else:
        print(Fore.RED+"Error launching game")
        driver_luna.save_screenshot(r'C:\Users\prebyada\Desktop\Screenshots\Error_in_launching_Germany.png')
        #increase/decrease launch time below
    time.sleep(game_wait)
        #change location of drive
    driver_luna.save_screenshot(r'C:\Users\prebyada\Desktop\Screenshots\OverideGermany_1.png')
    image = cv2.imread("C:\\Users\\prebyada\\Desktop\\Screenshots\\OverideGermany_1.png", 0)
    if cv2.countNonZero(image) == 0:
        print("Image is black")
    else:
        print("Colored image")
    send_keys('+{TAB}')
    time.sleep(10)
    send_keys("{VK_DOWN}")
    time.sleep(4)
    send_keys("{ENTER}")
    time.sleep(4)
    send_keys("{ENTER}")
    time.sleep(2)
    driver_luna.close()

    break